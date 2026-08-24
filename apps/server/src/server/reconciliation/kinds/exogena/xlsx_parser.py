"""Reads a DIAN exogena report straight out of its spreadsheet.

No AI in this path, deliberately. The file is a machine-generated table with a
fixed shape; a multimodal model would cost money per run, vary between runs,
and can silently misread a digit in a figure the entire report then rests on.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass

from openpyxl import load_workbook

from server.reconciliation.core.kind import SourceContent, SourceNotRecognized
from server.reconciliation.kinds.exogena.concepts import concept_code_in, concept_id_for
from server.shared import AccountRef, FactRole, FinancialFact, Money, Period, TaxId

logger = logging.getLogger(__name__)

XLSX_MEDIA_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }
)

# Column order under the `Persona que reporta | Información reportada` banner.
_REPORTER_TAX_ID, _REPORTER_NAME = 0, 1
_SUBJECT_TAX_ID, _SUBJECT_NAME = 2, 3
_DETAIL, _VALUE, _USAGE, _EXTRA = 4, 5, 6, 7

_HEADER_MARKERS = ("nit", "detalle", "valor")
_ACCOUNT = re.compile(r"n[uú]mero\s+de\s+cuenta\s*/?\s*documento\s*:\s*([0-9][0-9\s-]*)", re.I)
_YEAR = re.compile(r"\b(19|20)\d{2}\b")
_EXTRA_PAIR = re.compile(r"([^:|]+):\s*([^|]*)")


class ExogenaParseError(ValueError):
    """The file is an exogena report and could not be read."""


@dataclass(frozen=True, slots=True)
class _Header:
    row_index: int
    columns: dict[str, int]


class ExogenaXlsxExtractor:
    """Turns the exogena workbook into spine facts.

    Implements `reconciliation.core.kind.FactExtractor`.
    """

    def extract(self, content: SourceContent) -> tuple[FinancialFact, ...]:
        workbook = load_workbook(io.BytesIO(content.data), data_only=True, read_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            # The DIAN's generator writes a dimension range narrower than the
            # data it then writes. In read-only mode openpyxl trusts that range
            # and stops early, which silently drops most of the report — the
            # kind of failure that returns a clean-looking result rather than
            # an error. Recomputing the extent from the cells is what makes the
            # row count real.
            sheet.reset_dimensions()
            rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

        header = _find_header(rows)
        period = content.period or Period.of_year(_find_year(rows, header.row_index))
        subject = content.subject_tax_id or _find_subject(rows, header.row_index)

        facts: list[FinancialFact] = []
        for offset, row in enumerate(rows[header.row_index + 1 :], start=header.row_index + 2):
            fact = _row_to_fact(row, offset, period, subject, content.source_id)
            if fact is not None:
                facts.append(fact)
        if not facts:
            raise ExogenaParseError("The exogena report has a header but no reported rows")
        logger.info("Parsed exogena report", extra={"facts": len(facts), "period": period.key})
        return tuple(facts)


def _row_to_fact(
    row: list[object],
    row_number: int,
    period: Period,
    subject: TaxId | None,
    source_id: str,
) -> FinancialFact | None:
    reporter = TaxId.parse(_cell(row, _REPORTER_TAX_ID))
    # The summary block above the data ("Tope 1 - Ingresos", ...) has no
    # reporting party. Those are the DIAN's own aggregates over the rows below,
    # not third-party claims, so they are not facts to reconcile against a
    # certificate and are skipped rather than reported as unmatched.
    if reporter is None:
        return None

    detail = _cell(row, _DETAIL)
    amount = Money.parse(_cell(row, _VALUE))
    if amount is None or not detail:
        return None

    extra = _cell(row, _EXTRA)
    account_match = _ACCOUNT.search(extra)
    extras = {"usage": _cell(row, _USAGE), "reported_name": _cell(row, _SUBJECT_NAME)}
    code = concept_code_in(detail)
    if code:
        extras["dian_concept_code"] = code
    extras.update(_parse_extras(extra))

    return FinancialFact(
        source_id=source_id,
        role=FactRole.SPINE,
        reporter_tax_id=reporter,
        reporter_name=_cell(row, _REPORTER_NAME),
        subject_tax_id=TaxId.parse(_cell(row, _SUBJECT_TAX_ID)) or subject,
        concept_id=concept_id_for(detail),
        period=period,
        amount=amount,
        account=AccountRef.parse(account_match.group(1).strip()) if account_match else None,
        detail=detail.strip(),
        locator=f"row {row_number}",
        extras={k: v for k, v in extras.items() if v},
    )


def _parse_extras(text: str) -> dict[str, str]:
    """Split the free-text `Información Adicional` column into key/value pairs.

    It arrives as `Clase de Tarjeta: *1* T | Placa: WGX75D`, which is regular
    enough to read without a model and irregular enough that anything it fails
    to split is simply left out rather than guessed at.
    """
    pairs: dict[str, str] = {}
    for segment in text.split("|"):
        match = _EXTRA_PAIR.match(segment.strip())
        if match is None:
            continue
        key = re.sub(r"[^a-z0-9]+", "_", match.group(1).strip().lower()).strip("_")
        value = match.group(2).strip()
        if key and value:
            pairs[key] = value
    return pairs


def _find_header(rows: list[list[object]]) -> _Header:
    for index, row in enumerate(rows):
        cells = [str(c).strip().lower() for c in row if c is not None]
        if all(any(marker == cell for cell in cells) for marker in _HEADER_MARKERS):
            return _Header(row_index=index, columns={})
    raise SourceNotRecognized(
        "Could not find the reported-rows header; this does not look like a DIAN exogena report"
    )


def _find_year(rows: list[list[object]], header_row: int) -> int:
    """Read the tax year from the preamble above the table.

    Anchored on the label rather than on a cell position, because the preamble
    also carries the run date and picking the wrong one would file an entire
    report under the wrong year.
    """
    for row in rows[:header_row]:
        cells = [str(c) for c in row if c is not None]
        joined = " ".join(cells).lower()
        if "ano al que se refiere" in _fold(joined) or "año al que se refiere" in joined:
            for cell in cells:
                match = _YEAR.search(cell)
                if match:
                    return int(match.group(0))
    raise ExogenaParseError("Could not determine the tax year of the exogena report")


def _find_subject(rows: list[list[object]], header_row: int) -> TaxId | None:
    for row in rows[:header_row]:
        cells = [c for c in row if c is not None]
        labels = [str(c).strip().lower() for c in cells]
        if any(
            label.startswith("identificacion") or label.startswith("identificación")
            for label in labels
        ):
            for cell in cells[1:]:
                parsed = TaxId.parse(cell)
                if parsed is not None and len(parsed.value) >= 6:
                    return parsed
    return None


def _fold(text: str) -> str:
    import unicodedata

    decomposed = unicodedata.normalize("NFKD", text)
    return "".join(c for c in decomposed if not unicodedata.combining(c))


def _cell(row: list[object], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    value = row[index]
    # openpyxl hands back every numeric cell as a float, and a NIT is a numeric
    # cell: `str(890903938.0)` is `"890903938.0"`, whose digits are
    # `8909039380`. That trailing zero made every reporting party of the
    # exogena a different party from the one its certificate names, so not one
    # row could ever pair with its evidence.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
